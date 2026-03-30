"""
Core views for Entebbe Airport Slotting System.
"""

from datetime import date, datetime, timedelta
from typing import Any
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Q, Count
from urllib.parse import urlencode

from core.models import (
    Airline, Airport, AircraftType, FlightRequest, ParkingStand, Gate,
    CheckInCounter, StandAllocation, GateAllocation, CheckInAllocation, DAY_MASK
)
from core.services.season import (
    get_current_season, get_summer_dates, get_winter_dates
)
from core.services.allocation import allocate_resources_for_date, allocate_resources_for_flight, time_subtract_minutes


# ─── Temporary System Data Seed (remove after use) ─────────────────────────────

def seed_system_data(request: HttpRequest):
    """One-time seed view - adds missing aircraft types and airports. Remove after use."""
    new_aircraft = [
        ('B735','Boeing 737-500','Boeing','narrow_body','C',False,132),
        ('B733','Boeing 737-300','Boeing','narrow_body','C',False,148),
        ('B733F','Boeing 737-300 Freighter','Boeing','narrow_body','C',False,0),
        ('B738F','Boeing 737-800 Freighter','Boeing','narrow_body','C',False,0),
        ('B39M','Boeing 737 MAX 9','Boeing','narrow_body','C',False,220),
        ('BCS3','Airbus A220-300','Airbus','narrow_body','C',False,140),
        ('A332','Airbus A330-200','Airbus','wide_body','E',True,250),
        ('B78X','Boeing 787-10 Dreamliner','Boeing','wide_body','E',True,330),
        ('A359','Airbus A350-900','Airbus','wide_body','E',True,315),
        ('A35K','Airbus A350-1000','Airbus','wide_body','E',True,366),
    ]
    
    new_airports = [
        ('HRE', 'FVHA', 'Harare', 'Zimbabwe'),
        ('LUN', 'FLKK', 'Lusaka', 'Zambia'),
        ('BJM', 'HBBA', 'Bujumbura', 'Burundi'),
        ('JRO', 'HTKJ', 'Kilimanjaro', 'Tanzania'),
        ('MBA', 'HKMO', 'Mombasa', 'Kenya'),
        ('MGQ', 'HCMM', 'Mogadishu', 'Somalia'),
        ('BOM', 'VABB', 'Mumbai', 'India'),
        ('LGW', 'EGKK', 'London', 'United Kingdom'),
        ('DAR', 'HTDA', 'Dar es Salaam', 'Tanzania'),
        ('JNB', 'FAOR', 'Johannesburg', 'South Africa'),
        ('FIH', 'FZAA', 'Kinshasa', 'DR Congo'),
    ]

    aircraft_results = {}
    for code,name,mfr,cat,sz,wb,pax in new_aircraft:
        _, created = AircraftType.objects.get_or_create(
            code=code,
            defaults={'name':name,'manufacturer':mfr,'category':cat,'size_code':sz,'is_wide_body':wb,'pax_capacity':pax}
        )
        aircraft_results[code] = 'created' if created else 'already existed'
        
    airport_results = {}
    for iata, icao, city, country in new_airports:
        _, created = Airport.objects.get_or_create(
            iata_code=iata,
            defaults={'icao_code': icao, 'city_name': city, 'country': country}
        )
        airport_results[iata] = 'created' if created else 'already existed'

    return JsonResponse({
        'total_aircraft': AircraftType.objects.count(),
        'aircraft_results': aircraft_results,
        'total_airports': Airport.objects.count(),
        'airport_results': airport_results
    })


# ─── Dashboard ────────────────────────────────────────────────────────────────

def dashboard(request: HttpRequest):
    season_name, year, season_start, season_end = get_current_season()

    # Counts
    summer_count = FlightRequest.objects.filter(season='summer').count()
    winter_count = FlightRequest.objects.filter(season='winter').count()
    total_count = summer_count + winter_count

    gates_count = Gate.objects.filter(is_active=True).count()
    stands_count = ParkingStand.objects.filter(is_active=True, parent_stand__isnull=True).count()

    # This year's seasons
    summer_start_2026, summer_end_2026 = get_summer_dates(2026)
    winter_start_2025, winter_end_2025_26 = get_winter_dates(2025)

    # Operational actions / summaries
    pending_count = FlightRequest.objects.filter(status='pending').count()
    airlines_with_requests = FlightRequest.objects.values_list('airline_id', flat=True).distinct()
    missing_airlines_count = Airline.objects.exclude(id__in=airlines_with_requests).count()

    context = {
        'current_season': season_name,
        'current_year': year,
        'season_start': season_start,
        'season_end': season_end,
        'summer_count': summer_count,
        'winter_count': winter_count,
        'total_count': total_count,
        'gates_count': gates_count,
        'stands_count': stands_count,
        'summer_start_2026': summer_start_2026,
        'summer_end_2026': summer_end_2026,
        'winter_start_2025': winter_start_2025,
        'winter_end_2025_26': winter_end_2025_26,
        'pending_count': pending_count,
        'missing_airlines_count': missing_airlines_count,
        'active_page': 'dashboard',
    }
    return render(request, 'dashboard.html', context)


def admin_dashboard(request: HttpRequest):
    """Admin panel for managing airlines, airports, and aircraft."""
    airlines_count = Airline.objects.count()
    airports_count = Airport.objects.count()
    aircraft_count = AircraftType.objects.count()
    
    context = {
        'airlines_count': airlines_count,
        'airports_count': airports_count,
        'aircraft_count': aircraft_count,
        'active_page': 'admin',
    }
    return render(request, 'admin_dashboard.html', context)


# ─── Flight Requests ──────────────────────────────────────────────────────────

DAY_ORDER = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']


def _days_matrix(mask: int) -> list[bool]:
    return [bool(mask & DAY_MASK[day]) for day in DAY_ORDER]


def _redirect_to_next_or_default(
    request: HttpRequest,
    default_url_name: str,
    **kwargs: Any,
) -> HttpResponse:
    next_url = request.POST.get('next', '').strip()
    if next_url:
        return redirect(next_url)
    return redirect(default_url_name, **kwargs)

def flights_list(request: HttpRequest):
    season_filter = request.GET.get('season', 'all')
    airline_filter = request.GET.get('airline', '').strip()
    aircraft_filter = request.GET.get('aircraft', '').strip()
    day_filter = request.GET.get('day', '').strip().lower()

    qs = FlightRequest.objects.select_related('airline', 'aircraft_type', 'origin', 'destination')

    if season_filter in ('summer', 'winter'):
        qs = qs.filter(season=season_filter)
    if airline_filter.isdigit():
        qs = qs.filter(airline_id=int(airline_filter))
    if aircraft_filter.isdigit():
        qs = qs.filter(aircraft_type_id=int(aircraft_filter))

    flights_data = []
    for flight in qs.order_by('-created_at'):
        if day_filter in DAY_MASK and not (flight.days_of_operation & DAY_MASK[day_filter]):
            continue
        flights_data.append({
            'flight': flight,
            'days_matrix': _days_matrix(flight.days_of_operation),
        })

    query_params = {}
    if airline_filter:
        query_params['airline'] = airline_filter
    if aircraft_filter:
        query_params['aircraft'] = aircraft_filter
    if day_filter in DAY_MASK:
        query_params['day'] = day_filter
    query_suffix = f"&{urlencode(query_params)}" if query_params else ''

    summer_count = FlightRequest.objects.filter(season='summer').count()
    winter_count = FlightRequest.objects.filter(season='winter').count()

    context = {
        'flight_rows': flights_data,
        'season_filter': season_filter,
        'airline_filter': int(airline_filter) if airline_filter.isdigit() else airline_filter,
        'aircraft_filter': int(aircraft_filter) if aircraft_filter.isdigit() else aircraft_filter,
        'day_filter': day_filter,
        'query_suffix': query_suffix,
        'airlines': Airline.objects.all().order_by('name'),
        'aircraft_types': AircraftType.objects.all().order_by('code'),
        'day_options': [
            ('sunday', 'Sunday'),
            ('monday', 'Monday'),
            ('tuesday', 'Tuesday'),
            ('wednesday', 'Wednesday'),
            ('thursday', 'Thursday'),
            ('friday', 'Friday'),
            ('saturday', 'Saturday'),
        ],
        'day_headers': ['S', 'M', 'T', 'W', 'T', 'F', 'S'],
        'summer_count': summer_count,
        'winter_count': winter_count,
        'total_count': summer_count + winter_count,
        'active_page': 'flights',
    }
    return render(request, 'flights/index.html', context)


def flight_new(request: HttpRequest):
    airlines = Airline.objects.all().order_by('name')
    aircraft_types = AircraftType.objects.all().order_by('code')
    airports = Airport.objects.all().order_by('iata_code')
    summer_start_2026, summer_end_2026 = get_summer_dates(2026)
    winter_start_2025, winter_end_2025 = get_winter_dates(2025)

    context = {
        'airlines': airlines,
        'aircraft_types': aircraft_types,
        'airports': airports,
        'current_year': date.today().year,
        'summer_start': summer_start_2026,
        'summer_end': summer_end_2026,
        'winter_start': winter_start_2025,
        'winter_end': winter_end_2025,
        'active_page': 'flights',
        'form_action': 'new',
        'days': ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
    }
    return render(request, 'flights/form.html', context)


def flight_create(request: HttpRequest):
    if request.method != 'POST':
        return redirect('flight_new')

    try:
        airline = get_object_or_404(Airline, id=request.POST.get('airline_id'))
        aircraft_type = get_object_or_404(AircraftType, id=request.POST.get('aircraft_type_id'))

        origin_id = request.POST.get('origin_id')
        destination_id = request.POST.get('destination_id')
        origin = Airport.objects.filter(id=origin_id).first() if origin_id else None
        destination = Airport.objects.filter(id=destination_id).first() if destination_id else None

        # Days of operation bitmask
        day_names = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
        days_mask = 0
        for day in day_names:
            if request.POST.get(f'day_{day}'):
                days_mask |= DAY_MASK[day]

        # Partial season dates
        valid_from_str = request.POST.get('valid_from')
        valid_to_str = request.POST.get('valid_to')
        valid_from = datetime.strptime(valid_from_str, '%Y-%m-%d').date() if valid_from_str else None
        valid_to = datetime.strptime(valid_to_str, '%Y-%m-%d').date() if valid_to_str else None

        arrival_time_str = request.POST.get('arrival_time')
        departure_time_str = request.POST.get('departure_time')
        arrival_time = datetime.strptime(arrival_time_str, '%H:%M').time() if arrival_time_str else None
        departure_time = datetime.strptime(departure_time_str, '%H:%M').time() if departure_time_str else None

        flight = FlightRequest.objects.create(
            airline=airline,
            arrival_flight_number=request.POST.get('arrival_flight_number', '').strip(),
            departure_flight_number=request.POST.get('departure_flight_number', '').strip(),
            aircraft_type=aircraft_type,
            operation_type=request.POST.get('operation_type', 'turnaround'),
            season=request.POST.get('season', 'summer'),
            year=int(request.POST.get('year', date.today().year)),
            arrival_time=arrival_time,
            departure_time=departure_time,
            origin=origin,
            destination=destination,
            valid_from=valid_from,
            valid_to=valid_to,
            days_of_operation=days_mask,
            ground_days=int(request.POST.get('ground_days', 0)),
            notes=request.POST.get('notes', ''),
        )

        # Auto-allocate resources for every operating date in the season
        alloc_results = allocate_resources_for_flight(flight)
        if alloc_results['total_dates'] == 0:
            messages.success(request, 'Flight request created. Add arrival/departure times to enable auto-allocation.')
        elif alloc_results['conflicts'] == 0:
            messages.success(
                request,
                f'Flight request created and resources allocated across '
                f'{alloc_results["allocated"]} operating date(s).'
            )
        else:
            messages.warning(
                request,
                f'Flight request created — {alloc_results["allocated"]} date(s) allocated, '
                f'{alloc_results["conflicts"]} conflict(s) detected. '
                f'Check the Schedule page to resolve conflicts.'
            )
        return redirect('flights_list')

    except Exception as e:
        messages.error(request, f'Error creating flight request: {str(e)}')
        return redirect('flight_new')


def flight_edit(request: HttpRequest, pk: int):
    flight = get_object_or_404(FlightRequest, pk=pk)
    airlines = Airline.objects.all().order_by('name')
    aircraft_types = AircraftType.objects.all().order_by('code')
    airports = Airport.objects.all().order_by('iata_code')
    summer_start_2026, summer_end_2026 = get_summer_dates(2026)
    winter_start_2025, winter_end_2025 = get_winter_dates(2025)

    context = {
        'flight': flight,
        'airlines': airlines,
        'aircraft_types': aircraft_types,
        'airports': airports,
        'current_year': flight.year,
        'summer_start': summer_start_2026,
        'summer_end': summer_end_2026,
        'winter_start': winter_start_2025,
        'winter_end': winter_end_2025,
        'active_page': 'flights',
        'form_action': 'edit',
        'days': ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday'],
        'day_mask': flight.days_of_operation,
        'day_mask_map': DAY_MASK,
        'next_url': request.GET.get('next', ''),
    }
    return render(request, 'flights/form.html', context)


def flight_update(request: HttpRequest, pk: int):
    if request.method != 'POST':
        return redirect('flight_edit', pk=pk)
    flight = get_object_or_404(FlightRequest, pk=pk)

    try:
        airline = get_object_or_404(Airline, id=request.POST.get('airline_id'))
        aircraft_type = get_object_or_404(AircraftType, id=request.POST.get('aircraft_type_id'))

        origin_id = request.POST.get('origin_id')
        destination_id = request.POST.get('destination_id')
        origin = Airport.objects.filter(id=origin_id).first() if origin_id else None
        destination = Airport.objects.filter(id=destination_id).first() if destination_id else None

        day_names = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
        days_mask = 0
        for day in day_names:
            if request.POST.get(f'day_{day}'):
                days_mask |= DAY_MASK[day]

        valid_from_str = request.POST.get('valid_from')
        valid_to_str = request.POST.get('valid_to')
        valid_from = datetime.strptime(valid_from_str, '%Y-%m-%d').date() if valid_from_str else None
        valid_to = datetime.strptime(valid_to_str, '%Y-%m-%d').date() if valid_to_str else None

        arrival_time_str = request.POST.get('arrival_time')
        departure_time_str = request.POST.get('departure_time')
        arrival_time = datetime.strptime(arrival_time_str, '%H:%M').time() if arrival_time_str else None
        departure_time = datetime.strptime(departure_time_str, '%H:%M').time() if departure_time_str else None

        flight.airline = airline
        flight.arrival_flight_number = request.POST.get('arrival_flight_number', '').strip()
        flight.departure_flight_number = request.POST.get('departure_flight_number', '').strip()
        flight.aircraft_type = aircraft_type
        flight.operation_type = request.POST.get('operation_type', 'turnaround')
        flight.season = request.POST.get('season', 'summer')
        flight.year = int(request.POST.get('year', date.today().year))
        flight.arrival_time = arrival_time
        flight.departure_time = departure_time
        flight.origin = origin
        flight.destination = destination
        flight.valid_from = valid_from
        flight.valid_to = valid_to
        flight.days_of_operation = days_mask
        flight.ground_days = int(request.POST.get('ground_days', 0))
        flight.notes = request.POST.get('notes', '')
        flight.status = 'pending'  # reset status when edited
        flight.save()

        # Clear old allocations before re-allocating
        StandAllocation.objects.filter(flight_request=flight).delete()
        GateAllocation.objects.filter(flight_request=flight).delete()
        CheckInAllocation.objects.filter(flight_request=flight).delete()

        # Re-allocate resources for every operating date
        alloc_results = allocate_resources_for_flight(flight)
        if alloc_results['total_dates'] == 0:
            messages.success(request, 'Flight request updated. Add arrival/departure times to enable auto-allocation.')
        elif alloc_results['conflicts'] == 0:
            messages.success(
                request,
                f'Flight request updated and resources re-allocated across '
                f'{alloc_results["allocated"]} operating date(s).'
            )
        else:
            messages.warning(
                request,
                f'Flight request updated — {alloc_results["allocated"]} date(s) allocated, '
                f'{alloc_results["conflicts"]} conflict(s) detected. '
                f'Check the Schedule page to resolve conflicts.'
            )
        return _redirect_to_next_or_default(request, 'flights_list')

    except Exception as e:
        messages.error(request, f'Error updating flight request: {str(e)}')
        return redirect('flight_edit', pk=pk)


@require_POST
def flight_delete(request: HttpRequest, pk: int):
    flight = get_object_or_404(FlightRequest, pk=pk)
    flight.delete()
    messages.success(request, f'Flight request deleted.')
    return _redirect_to_next_or_default(request, 'flights_list')


@require_POST
def flight_approve(request: HttpRequest, pk: int):
    flight = get_object_or_404(FlightRequest, pk=pk)
    flight.status = 'allocated'
    flight.save(update_fields=['status', 'updated_at'])
    messages.success(request, f'Flight request {flight} approved.')
    return _redirect_to_next_or_default(request, 'flights_list')


@require_POST
def flight_reject(request: HttpRequest, pk: int):
    flight = get_object_or_404(FlightRequest, pk=pk)
    flight.status = 'cancelled'
    flight.save(update_fields=['status', 'updated_at'])
    StandAllocation.objects.filter(flight_request=flight).delete()
    GateAllocation.objects.filter(flight_request=flight).delete()
    CheckInAllocation.objects.filter(flight_request=flight).delete()
    messages.info(request, f'Flight request {flight} rejected.')
    return _redirect_to_next_or_default(request, 'flights_list')


# ─── Daily Schedule ───────────────────────────────────────────────────────────

def schedule_view(request: HttpRequest):
    date_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()

    from core.services.season import get_season_for_date
    season, year = get_season_for_date(selected_date)

    # Get all flight requests for this season
    all_requests = FlightRequest.objects.filter(
        season=season, year=year
    ).select_related('airline', 'aircraft_type', 'origin', 'destination')

    # Filter to those that operate on this specific date
    daily_flights = [f for f in all_requests if f.operates_on_date(selected_date)]

    # Get allocations for this date
    stand_allocs = {
        int(sa.flight_request.pk): sa
        for sa in StandAllocation.objects.filter(date=selected_date).select_related('stand', 'flight_request')
    }
    gate_allocs = {
        int(ga.flight_request.pk): ga
        for ga in GateAllocation.objects.filter(date=selected_date).select_related('gate', 'flight_request')
    }
    checkin_allocs = {
        int(ca.flight_request.pk): ca
        for ca in CheckInAllocation.objects.filter(date=selected_date).select_related('flight_request')
    }

    # Pre-fetch handlers for the daily flights
    from core.models import GroundHandler, AirlineGatePreference, AirlineStandPreference
    handlers_map = {}
    for handler in GroundHandler.objects.prefetch_related('airlines').all():
        for airline in handler.airlines.all():
            handlers_map[int(airline.pk)] = handler

    stand_prefs = {}
    for p in AirlineStandPreference.objects.select_related('destination').all():
        airline_pk = int(p.airline.pk)
        if airline_pk not in stand_prefs:
            stand_prefs[airline_pk] = []
        stand_prefs[airline_pk].append(p)
        
    gate_prefs = {}
    for p in AirlineGatePreference.objects.select_related('preferred_gate', 'destination').all():
        airline_pk = int(p.airline.pk)
        if airline_pk not in gate_prefs:
            gate_prefs[airline_pk] = []
        gate_prefs[airline_pk].append(p)

    # Build display rows
    flight_rows = []
    for f in sorted(daily_flights, key=lambda x: x.arrival_time or x.departure_time or datetime.min.time()):
        flight_pk = int(f.pk)
        airline_pk = int(f.airline.pk)
        dest_pk = int(f.destination.pk) if f.destination else None
        handler = handlers_map.get(airline_pk)
        
        warnings = []
        stand = stand_allocs.get(flight_pk)
        gate = gate_allocs.get(flight_pk)
        
        # Stand warning
        for pref in stand_prefs.get(airline_pk, []):
            pref_dest_pk = int(pref.destination.pk) if pref.destination else None
            if pref_dest_pk is None or pref_dest_pk == dest_pk:
                if pref.requires_bridge and stand and not stand.stand.has_boarding_bridge:
                    warnings.append(f"Requires bridge stand (currently on Stand {stand.stand.stand_number})")
                break
                
        # Gate warning
        for pref in gate_prefs.get(airline_pk, []):
            pref_dest_pk = int(pref.destination.pk) if pref.destination else None
            preferred_gate_pk = int(pref.preferred_gate.pk)
            if pref_dest_pk is None or pref_dest_pk == dest_pk:
                if gate and int(gate.gate.pk) != preferred_gate_pk:
                    warnings.append(f"Preferred gate is {pref.preferred_gate.gate_number} (currently on Gate {gate.gate.gate_number})")
                break
                
        flight_rows.append({
            'flight': f,
            'stand': stand_allocs.get(flight_pk),
            'gate': gate_allocs.get(flight_pk),
            'checkin': checkin_allocs.get(flight_pk),
            'allocated': flight_pk in stand_allocs or flight_pk in gate_allocs or flight_pk in checkin_allocs,
            'handler': handler,
        })

    allocated_count = sum(1 for r in flight_rows if r['allocated'])
    conflict_count = sum(1 for r in flight_rows if r['flight'].status == 'conflict')

    context = {
        'selected_date': selected_date,
        'date_str': date_str,
        'schedule_return_url': f'/schedule/?date={date_str}',
        'flight_rows': flight_rows,
        'total_flights': len(flight_rows),
        'allocated_count': allocated_count,
        'conflict_count': conflict_count,
        'active_page': 'schedule',
    }
    return render(request, 'schedule.html', context)


@require_POST
def schedule_allocate(request: HttpRequest):
    date_str = request.POST.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        alloc_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        alloc_date = date.today()

    results = allocate_resources_for_date(alloc_date)
    messages.success(
        request,
        f"Allocation complete: {results['allocated']} allocated, "
        f"{results['conflicts']} conflicts, {results['skipped']} skipped."
    )
    return redirect(f'/schedule/?date={date_str}')


@require_POST
def schedule_clear(request: HttpRequest):
    """Clear all allocations for a specific date."""
    date_str = request.POST.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        clear_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        clear_date = date.today()

    # Collect affected flight IDs before deleting
    affected_ids = set(
        StandAllocation.objects.filter(date=clear_date).values_list('flight_request_id', flat=True)
    ) | set(
        GateAllocation.objects.filter(date=clear_date).values_list('flight_request_id', flat=True)
    ) | set(
        CheckInAllocation.objects.filter(date=clear_date).values_list('flight_request_id', flat=True)
    )

    StandAllocation.objects.filter(date=clear_date).delete()
    GateAllocation.objects.filter(date=clear_date).delete()
    CheckInAllocation.objects.filter(date=clear_date).delete()

    # Only reset status for flights with NO remaining allocations on any other date
    for fid in affected_ids:
        if not StandAllocation.objects.filter(flight_request_id=fid).exists():
            FlightRequest.objects.filter(id=fid, status='allocated').update(status='pending')

    messages.info(request, f"All allocations for {clear_date.strftime('%d/%m/%Y')} cleared.")
    return redirect(f'/schedule/?date={date_str}')


def schedule_allocate_manual(request: HttpRequest, flight_id: int):
    """Show available resources for manual allocation."""
    from core.services.allocation import (
        time_subtract_minutes, times_overlap
    )
    
    flight = get_object_or_404(FlightRequest, pk=flight_id)
    date_str = request.GET.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()
    
    if not flight.operates_on_date(selected_date):
        messages.error(request, 'Flight does not operate on selected date.')
        return redirect(f'/schedule/?date={date_str}')
    
    arrival = flight.arrival_time or flight.departure_time
    departure = flight.departure_time or flight.arrival_time
    
    # Find hard blocked stands and gates
    from core.models import AirlineGatePreference
    hard_blocked_gates = []
    hard_blocked_stands = []
    
    gate_blocks = AirlineGatePreference.objects.filter(is_hard_block=True)
    for gb in gate_blocks:
        blocking_flights = FlightRequest.objects.filter(airline=gb.airline)
        if gb.destination:
            blocking_flights = blocking_flights.filter(destination=gb.destination)
        for bf in blocking_flights:
            if int(bf.pk) != int(flight.pk) and bf.operates_on_date(selected_date):
                arr = bf.arrival_time or bf.departure_time
                dep = bf.departure_time or bf.arrival_time
                if bf.operation_type != 'arrival' and bf.departure_time:
                    g_close = time_subtract_minutes(bf.departure_time, 15)
                    g_open = time_subtract_minutes(bf.departure_time, 45)
                    hard_blocked_gates.append((int(gb.preferred_gate.pk), g_open, g_close))
                if arr and dep and gb.preferred_gate.connected_stand_id:
                    hard_blocked_stands.append((gb.preferred_gate.connected_stand_id, arr, dep))

    # Get available stands (as dict options)
    existing_stand_allocs = StandAllocation.objects.filter(date=selected_date).select_related('flight_request', 'flight_request__airline')
    all_stands = ParkingStand.objects.filter(
        is_active=True, parent_stand__isnull=True
    ).filter(Q(size_code__in=['D', 'E', 'F']) | Q(size_code__lte=flight.aircraft_type.size_code))
    
    available_stands = []
    for stand in all_stands:
        if stand.can_accommodate(flight.aircraft_type):
            is_available = True
            conflict_msg = ""
            
            # Check hard blocked first
            for sid, sstart, send in hard_blocked_stands:
                if sid == int(stand.pk) and times_overlap(arrival, departure, sstart, send):
                    is_available = False
                    conflict_msg = "Hard-blocked"
                    break
            
            # Check existing allocations
            if is_available:
                for alloc in existing_stand_allocs:
                    if int(alloc.stand.pk) == int(stand.pk) and times_overlap(arrival, departure, alloc.start_time, alloc.end_time):
                        is_available = False
                        f_disp = alloc.flight_request.display_flight_numbers
                        conflict_msg = f"In use by {f_disp} ({alloc.start_time.strftime('%H:%M')} – {alloc.end_time.strftime('%H:%M')})"
                        break
                        
            available_stands.append({
                'id': int(stand.pk),
                'name': f"Stand {stand.stand_number} (Code {stand.size_code}{', Bridge' if stand.has_boarding_bridge else ''})",
                'is_available': is_available,
                'conflict_msg': conflict_msg
            })
    
    # Get available gates
    available_gates = []
    if flight.operation_type != 'arrival':
        departure_time = flight.departure_time
        if departure_time is None:
            departure_time = flight.arrival_time
        if departure_time is not None:
            gate_close = time_subtract_minutes(departure_time, 15)
            gate_open = time_subtract_minutes(departure_time, 45)
        
            existing_gate_allocs = GateAllocation.objects.filter(date=selected_date).select_related('flight_request', 'flight_request__airline')
            all_gates = Gate.objects.filter(is_active=True)
            
            for gate in all_gates:
                is_available = True
                conflict_msg = ""
                
                for gid, gstart, gend in hard_blocked_gates:
                    if gid == int(gate.pk) and times_overlap(gate_open, gate_close, gstart, gend):
                        is_available = False
                        conflict_msg = "Hard-blocked"
                        break
                
                if is_available:
                    for alloc in existing_gate_allocs:
                        if int(alloc.gate.pk) == int(gate.pk) and times_overlap(gate_open, gate_close, alloc.start_time, alloc.end_time):
                            is_available = False
                            f_disp = alloc.flight_request.display_flight_numbers
                            conflict_msg = f"In use by {f_disp} ({alloc.start_time.strftime('%H:%M')} – {alloc.end_time.strftime('%H:%M')})"
                            break
                            
                available_gates.append({
                    'id': int(gate.pk),
                    'name': f"Gate {gate.gate_number}{' (Bridge)' if gate.has_boarding_bridge else ''}",
                    'is_available': is_available,
                    'conflict_msg': conflict_msg
                })
    
    # Get available counters
    available_counters = []
    if flight.operation_type != 'arrival':
        departure_time = flight.departure_time
        if departure_time is not None:
            checkin_close = time_subtract_minutes(departure_time, 60)
            checkin_open = time_subtract_minutes(checkin_close, flight.checkin_duration_hours * 60)
            
            existing_checkin_allocs = list(CheckInAllocation.objects.filter(date=selected_date).select_related('flight_request', 'flight_request__airline'))
            num_needed = flight.min_counters
            
            for start in range(1, 24 - num_needed + 1):
                end = start + num_needed - 1
                is_available = True
                conflict_msg = ""
                
                for alloc in existing_checkin_allocs:
                    for c in range(start, end + 1):
                        if alloc.counter_from <= c <= alloc.counter_to:
                            if times_overlap(checkin_open, checkin_close, alloc.start_time, alloc.end_time):
                                is_available = False
                                f_disp = alloc.flight_request.display_flight_numbers
                                conflict_msg = f"In use by {f_disp} ({alloc.start_time.strftime('%H:%M')} – {alloc.end_time.strftime('%H:%M')})"
                                break
                    if not is_available:
                        break
                        
                available_counters.append({
                    'id': f"{start}-{end}",
                    'name': f"Counters {start}–{end} ({end - start + 1} counters)",
                    'is_available': is_available,
                    'conflict_msg': conflict_msg
                })
    
    context = {
        'flight': flight,
        'date_str': date_str,
        'selected_date': selected_date,
        'available_stands': available_stands,
        'available_gates': available_gates,
        'available_counters': available_counters,
    }
    return render(request, 'schedule_allocate_manual.html', context)


@require_POST
def schedule_allocate_manual_submit(request: HttpRequest, flight_id: int):
    """Process manual resource allocation."""
    flight = get_object_or_404(FlightRequest, pk=flight_id)
    date_str = request.POST.get('date', date.today().strftime('%Y-%m-%d'))
    try:
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = date.today()
    
    stand_id = request.POST.get('stand_id')
    gate_id = request.POST.get('gate_id')
    counter_range = request.POST.get('counter_range')
    
    try:
        # Allocate stand
        if stand_id:
            stand = get_object_or_404(ParkingStand, pk=stand_id)
            arrival = flight.arrival_time or flight.departure_time
            departure = flight.departure_time or flight.arrival_time
            StandAllocation.objects.update_or_create(
                flight_request=flight,
                date=selected_date,
                defaults={
                    'stand': stand,
                    'start_time': arrival,
                    'end_time': departure,
                }
            )
        
        # Allocate gate
        if gate_id and flight.operation_type != 'arrival':
            gate = get_object_or_404(Gate, pk=gate_id)
            departure_time = flight.departure_time
            if departure_time is not None:
                gate_close = time_subtract_minutes(departure_time, 15)
                gate_open = time_subtract_minutes(departure_time, 45)
                GateAllocation.objects.update_or_create(
                    flight_request=flight,
                    date=selected_date,
                    defaults={
                        'gate': gate,
                        'start_time': gate_open,
                        'end_time': gate_close,
                    }
                )
        
        # Allocate check-in counters
        if counter_range and flight.operation_type != 'arrival':
            counter_from, counter_to = map(int, counter_range.split('-'))
            departure_time = flight.departure_time
            if departure_time is not None:
                checkin_close = time_subtract_minutes(departure_time, 60)
                checkin_open = time_subtract_minutes(checkin_close, flight.checkin_duration_hours * 60)
                CheckInAllocation.objects.update_or_create(
                    flight_request=flight,
                    date=selected_date,
                    defaults={
                        'counter_from': counter_from,
                        'counter_to': counter_to,
                        'start_time': checkin_open,
                        'end_time': checkin_close,
                    }
                )
        
        flight.status = 'allocated'
        flight.save(update_fields=['status', 'updated_at'])
        messages.success(request, f'Resources allocated for {flight}.')
    except Exception as e:
        messages.error(request, f'Error allocating resources: {str(e)}')
    
    return redirect(f'/schedule/?date={date_str}')




# ─── Season Resource Allocation ────────────────────────────────────────────────

def season_allocations_view(request: HttpRequest):
    """
    Show all flights for a season with their resource allocations.
    Allow bulk assignment by airline and check for conflicts.
    """
    from core.services.season import get_current_season
    
    season_filter = request.GET.get('season', 'all')
    airline_filter = request.GET.get('airline', '').strip()
    
    season_name, year, season_start, season_end = get_current_season()
    
    # Get flights for the season
    qs = FlightRequest.objects.select_related('airline', 'aircraft_type', 'origin', 'destination')
    
    if season_filter in ('summer', 'winter'):
        qs = qs.filter(season=season_filter)
    else:
        # Get both summer and winter
        qs = qs.filter(year=year)
    
    if airline_filter.isdigit():
        qs = qs.filter(airline_id=int(airline_filter))
    
    qs = qs.order_by('airline__name', 'arrival_time', 'departure_time')
    
    # Pre-fetch handlers for the filtered airlines
    from core.models import GroundHandler, AirlineGatePreference, AirlineStandPreference
    handlers_map = {}
    for handler in GroundHandler.objects.prefetch_related('airlines').all():
        for airline in handler.airlines.all():
            handlers_map[int(airline.pk)] = handler

    # Pre-fetch preferences for warnings
    stand_prefs = {}
    for p in AirlineStandPreference.objects.select_related('destination').all():
        airline_pk = int(p.airline.pk)
        if airline_pk not in stand_prefs:
            stand_prefs[airline_pk] = []
        stand_prefs[airline_pk].append(p)
        
    gate_prefs = {}
    for p in AirlineGatePreference.objects.select_related('preferred_gate', 'destination').all():
        airline_pk = int(p.airline.pk)
        if airline_pk not in gate_prefs:
            gate_prefs[airline_pk] = []
        gate_prefs[airline_pk].append(p)
    
    # Pre-fetch all allocations to calculate specific conflict reasons efficiently
    flight_ids = [int(f.pk) for f in qs]
    
    stand_dates_qs = StandAllocation.objects.filter(flight_request_id__in=flight_ids).values('flight_request_id', 'date')
    gate_dates_qs = GateAllocation.objects.filter(flight_request_id__in=flight_ids).values('flight_request_id', 'date')
    checkin_dates_qs = CheckInAllocation.objects.filter(flight_request_id__in=flight_ids).values('flight_request_id', 'date')
    
    flight_stand_dates = {}
    flight_gate_dates = {}
    flight_checkin_dates = {}
    
    for sa in stand_dates_qs:
        flight_stand_dates.setdefault(sa['flight_request_id'], set()).add(sa['date'])
    for ga in gate_dates_qs:
        flight_gate_dates.setdefault(ga['flight_request_id'], set()).add(ga['date'])
    for ca in checkin_dates_qs:
        flight_checkin_dates.setdefault(ca['flight_request_id'], set()).add(ca['date'])
    
    # Build flight rows with allocations
    flight_rows = []
    for flight in qs:
        # Get the most common allocation (season-level)
        stand_allocs = StandAllocation.objects.filter(flight_request=flight).values('stand_id').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        gate_allocs = GateAllocation.objects.filter(flight_request=flight).values('gate_id').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        checkin_allocs = CheckInAllocation.objects.filter(flight_request=flight).values(
            'counter_from', 'counter_to'
        ).annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        # Check for conflicts on any date
        has_conflict = FlightRequest.objects.filter(id=int(flight.pk), status='conflict').exists()
        
        stand = None
        gate = None
        checkin = None
        
        if stand_allocs:
            try:
                stand = ParkingStand.objects.get(id=stand_allocs['stand_id'])
            except:
                pass
        
        if gate_allocs:
            try:
                gate = Gate.objects.get(id=gate_allocs['gate_id'])
            except:
                pass
        
        if checkin_allocs:
            checkin = f"C{checkin_allocs['counter_from']}–{checkin_allocs['counter_to']}"
        
        airline_pk = int(flight.airline.pk)
        dest_pk = int(flight.destination.pk) if flight.destination else None
        handler = handlers_map.get(airline_pk)
        
        # Check warnings
        warnings = []
        
        # Stand warning: requires bridge but got non-bridge
        flight_stand_prefs = stand_prefs.get(airline_pk, [])
        for pref in flight_stand_prefs:
            pref_dest_pk = int(pref.destination.pk) if pref.destination else None
            if pref_dest_pk is None or pref_dest_pk == dest_pk:
                if pref.requires_bridge and stand and not stand.has_boarding_bridge:
                    warnings.append(f"Requires bridge stand (currently on Stand {stand.stand_number})")
                break
                
        # Gate warning
        flight_gate_prefs = gate_prefs.get(airline_pk, [])
        for pref in flight_gate_prefs:
            pref_dest_pk = int(pref.destination.pk) if pref.destination else None
            if pref_dest_pk is None or pref_dest_pk == dest_pk:
                if gate and int(gate.pk) != int(pref.preferred_gate.pk):
                    warnings.append(f"Preferred gate is {pref.preferred_gate.gate_number} (currently on Gate {gate.gate_number})")
                break
                
        conflict_reasons = []
        if has_conflict:
            range_start = flight.valid_from or season_start
            range_end = flight.valid_to or season_end
            
            flight_pk = int(flight.pk)
            s_dates = flight_stand_dates.get(flight_pk, set())
            g_dates = flight_gate_dates.get(flight_pk, set())
            c_dates = flight_checkin_dates.get(flight_pk, set())
            
            current = range_start
            has_stand_missing = False
            has_gate_missing = False
            has_checkin_missing = False
            
            while current <= range_end:
                if flight.operates_on_date(current):
                    if current not in s_dates:
                        has_stand_missing = True
                    
                    if flight.operation_type != 'arrival':
                        dep_date = current + timedelta(days=flight.ground_days)
                        if dep_date not in g_dates:
                            has_gate_missing = True
                        if dep_date not in c_dates:
                            has_checkin_missing = True
                current += timedelta(days=1)
                
            if has_stand_missing:
                conflict_reasons.append("Stand")
            if has_gate_missing:
                conflict_reasons.append("Gate")
            if has_checkin_missing:
                conflict_reasons.append("Check-in")
        
        flight_rows.append({
            'flight': flight,
            'stand': stand,
            'gate': gate,
            'checkin': checkin,
            'has_conflict': has_conflict,
            'conflict_reasons': ", ".join(conflict_reasons),
            'allocated': stand or gate or checkin,
            'handler': handler,
            'warnings': warnings,
        })
    
    # Get counts by season
    summer_count = FlightRequest.objects.filter(season='summer', year=year).count()
    winter_count = FlightRequest.objects.filter(season='winter', year=year).count()
    
    # Get all available stands and gates for dropdown
    _all_stands = ParkingStand.objects.filter(is_active=True, parent_stand__isnull=True).order_by('stand_number')
    _all_gates = Gate.objects.filter(is_active=True).order_by('gate_number')
    
    context = {
        'flight_rows': flight_rows,
        'season_filter': season_filter,
        'airline_filter': int(airline_filter) if airline_filter.isdigit() else airline_filter,
        'airlines': Airline.objects.all().order_by('name'),
        'summer_count': summer_count,
        'winter_count': winter_count,
        'total_count': summer_count + winter_count,
        'current_season': season_name,
        'current_year': year,
        'season_start': season_start,
        'season_end': season_end,
        'active_page': 'allocations',
    }
    return render(request, 'allocations.html', context)


def flight_allocate_season(request: HttpRequest, flight_id: int):
    """
    Show detailed allocation page for a single flight across the entire season.
    """
    flight = get_object_or_404(FlightRequest.objects.select_related('airline', 'aircraft_type', 'origin', 'destination'), pk=flight_id)
    from core.services.season import get_season_dates
    
    # Get season dates
    if flight.season == 'summer':
        season_start, season_end = get_season_dates('summer', flight.year)
    else:
        season_start, season_end = get_season_dates('winter', flight.year)
    
    # Get current allocations
    stand_allocs = StandAllocation.objects.filter(flight_request=flight).select_related('stand')
    gate_allocs = GateAllocation.objects.filter(flight_request=flight).select_related('gate')
    checkin_allocs = CheckInAllocation.objects.filter(flight_request=flight)
    
    # Build allocation map by date
    stand_map = {alloc.date: alloc.stand for alloc in stand_allocs}
    gate_map = {alloc.date: alloc.gate for alloc in gate_allocs}
    checkin_map = {alloc.date: f"{alloc.counter_from}-{alloc.counter_to}" for alloc in checkin_allocs}
    
    # Get all available resources
    all_stands = ParkingStand.objects.filter(is_active=True, parent_stand__isnull=True).order_by('apron', 'stand_number')
    all_gates = Gate.objects.filter(is_active=True).order_by('gate_number')
    
    # Organize stands by apron
    apron1_stands = all_stands.filter(apron='apron1')
    apron1ext_stands = all_stands.filter(apron='apron1ext')
    
    # Get operating dates
    operating_dates = []
    current_date = season_start
    while current_date <= season_end:
        if flight.operates_on_date(current_date):
            operating_dates.append(current_date)
        current_date += timedelta(days=1)
    
    days_count = len(operating_dates)
    
    context = {
        'flight': flight,
        'season_start': season_start,
        'season_end': season_end,
        'operating_dates': operating_dates,
        'days_count': days_count,
        'stand_map': stand_map,
        'gate_map': gate_map,
        'checkin_map': checkin_map,
        'apron1_stands': apron1_stands,
        'apron1ext_stands': apron1ext_stands,
        'all_gates': all_gates,
        'active_page': 'allocations',
        'season_filter': request.GET.get('season', flight.season),
    }
    return render(request, 'flight_allocate_season.html', context)


@require_POST
def flight_allocate_season_submit(request: HttpRequest, flight_id: int):
    """
    Save resource allocations for a flight across the entire season.
    """
    flight = get_object_or_404(FlightRequest, pk=flight_id)
    from core.services.season import get_season_dates
    from core.services.allocation import time_subtract_minutes
    
    stand_id = request.POST.get('stand_id')
    gate_id = request.POST.get('gate_id')
    counter_from = request.POST.get('counter_from')
    counter_to = request.POST.get('counter_to')
    season_filter = request.POST.get('season', flight.season)
    
    # Convert to integers
    try:
        stand_id = int(stand_id) if stand_id and stand_id.strip() and stand_id != '' else None
    except ValueError:
        stand_id = None
    
    try:
        gate_id = int(gate_id) if gate_id and gate_id.strip() and gate_id != '' else None
    except ValueError:
        gate_id = None
    
    try:
        counter_from = int(counter_from) if counter_from and counter_from.strip() else None
        counter_to = int(counter_to) if counter_to and counter_to.strip() else None
    except ValueError:
        counter_from = None
        counter_to = None
    
    # Get season dates
    if flight.season == 'summer':
        season_start, season_end = get_season_dates('summer', flight.year)
    else:
        season_start, season_end = get_season_dates('winter', flight.year)
    
    try:
        resource_assigned = False
        # wipe existing allocations first so we don't accumulate stale rows
        StandAllocation.objects.filter(flight_request=flight).delete()
        GateAllocation.objects.filter(flight_request=flight).delete()
        CheckInAllocation.objects.filter(flight_request=flight).delete()
        
        # Assign stand for all operating dates
        if stand_id:
            stand = ParkingStand.objects.get(id=stand_id)
            current_date = season_start
            while current_date <= season_end:
                if flight.operates_on_date(current_date):
                    arrival = flight.arrival_time or flight.departure_time
                    departure = flight.departure_time or flight.arrival_time
                    if arrival and departure:
                        StandAllocation.objects.update_or_create(
                            flight_request=flight,
                            date=current_date,
                            defaults={
                                'stand': stand,
                                'start_time': arrival,
                                'end_time': departure,
                            }
                        )
                        resource_assigned = True
                current_date += timedelta(days=1)
        
        # Assign gate for all operating dates
        if gate_id and flight.operation_type != 'arrival':
            gate = Gate.objects.get(id=gate_id)
            current_date = season_start
            while current_date <= season_end:
                if flight.operates_on_date(current_date):
                    departure_time = flight.departure_time
                    if departure_time:
                        gate_close = time_subtract_minutes(departure_time, 15)
                        gate_open = time_subtract_minutes(departure_time, 45)
                        GateAllocation.objects.update_or_create(
                            flight_request=flight,
                            date=current_date,
                            defaults={
                                'gate': gate,
                                'start_time': gate_open,
                                'end_time': gate_close,
                            }
                        )
                        resource_assigned = True
                current_date += timedelta(days=1)
        
        # Assign check-in counters for all operating dates
        if counter_from and counter_to and flight.operation_type != 'arrival':
            current_date = season_start
            while current_date <= season_end:
                if flight.operates_on_date(current_date):
                    departure_time = flight.departure_time
                    if departure_time:
                        checkin_close = time_subtract_minutes(departure_time, 40)
                        checkin_open = time_subtract_minutes(departure_time, 180)
                        CheckInAllocation.objects.update_or_create(
                            flight_request=flight,
                            date=current_date,
                            defaults={
                                'counter_from': counter_from,
                                'counter_to': counter_to,
                                'start_time': checkin_open,
                                'end_time': checkin_close,
                            }
                        )
                        resource_assigned = True
                current_date += timedelta(days=1)
        
        # Update flight status
        if resource_assigned:
            flight.status = 'allocated'
            flight.save()
            messages.success(request, f'Resources assigned to {flight.arrival_flight_number or flight.departure_flight_number} for entire season.')
        else:
            messages.warning(request, 'No resources were assigned.')
    
    except Exception as e:
        messages.error(request, f'Error assigning resources: {str(e)}')
    
    return redirect(f'/allocations/?season={season_filter}')


def flight_resolve_conflict(request: HttpRequest, flight_id: int):
    """
    Shows a UI to resolve conflicts by suggesting alternative flight times.
    """
    flight = get_object_or_404(FlightRequest, pk=flight_id)
    season_filter = request.GET.get('season', flight.season)
    
    from core.services.conflict_resolution import find_alternative_slots
    
    # Check if there are actual conflicts before doing expensive search
    if flight.status != 'conflict':
        messages.info(request, "This flight does not have any conflicts to resolve.")
        return redirect(f'/allocations/?season={season_filter}')
        
    # Generate alternative timings
    suggestions = find_alternative_slots(flight, max_hours_search=4, interval_mins=15)
    
    context = {
        'flight': flight,
        'season_filter': season_filter,
        'suggestions': suggestions,
        'has_suggestions': len(suggestions) > 0,
        'active_page': 'allocations'
    }
    return render(request, 'flight_resolve.html', context)


@require_POST
def flight_apply_resolution(request: HttpRequest, flight_id: int):
    """
    Applies an alternative time schedule for the flight or rejects it.
    """
    flight = get_object_or_404(FlightRequest, pk=flight_id)
    season_filter = request.POST.get('season', flight.season)
    action = request.POST.get('action')
    
    if action == 'reject':
        flight.status = 'cancelled'
        flight.save()
        StandAllocation.objects.filter(flight_request=flight).delete()
        GateAllocation.objects.filter(flight_request=flight).delete()
        CheckInAllocation.objects.filter(flight_request=flight).delete()
        messages.success(request, f"Flight {flight.display_flight_numbers} has been rejected/cancelled.")
        return redirect(f'/allocations/?season={season_filter}')
        
    elif action == 'apply_time':
        arrival_str = request.POST.get('arrival_time')
        departure_str = request.POST.get('departure_time')
        
        try:
            if arrival_str and arrival_str != 'None':
                flight.arrival_time = datetime.strptime(arrival_str, '%H:%M:%S').time()
            if departure_str and departure_str != 'None':
                flight.departure_time = datetime.strptime(departure_str, '%H:%M:%S').time()
                
            flight.status = 'pending'
            flight.save()
            
            # Clear old and reallocate
            StandAllocation.objects.filter(flight_request=flight).delete()
            GateAllocation.objects.filter(flight_request=flight).delete()
            CheckInAllocation.objects.filter(flight_request=flight).delete()
            
            results = allocate_resources_for_flight(flight)
            
            if flight.status == 'allocated':
                messages.success(request, f"Successfully rescheduled and allocated flight to new times.")
            else:
                messages.warning(request, f"Schedule updated, but some conflicts remain ({results['conflicts']} conflicts).")
                
        except Exception as e:
            messages.error(request, f"Error updating flight times: {str(e)}")
            
    return redirect(f'/allocations/?season={season_filter}')


@require_POST
def season_allocations_auto(request: HttpRequest):
    """
    Auto-allocate resources to selected flights using the allocation service.

    This view is used when the user selects one or more flights and clicks the
    "Auto Allocate" button.  Previously the implementation simply looped over
    the requested dates and called the allocation helpers, leaving any existing
    allocations in place.  That behaviour made it impossible to "refresh" an
    allocation: once a stand/gate/check‑in had been assigned the helper would
    never relinquish it, causing the system to repeatedly reuse the same
    records and never explore alternative configurations.

    The new behaviour below does the following:

    * wipes all existing allocations for each flight at the start, ensuring we
      always start from a clean slate; the resulting assignments may be
      identical to the previous ones, but nothing is preserved implicitly.
    * passes ``shuffle=True`` to the allocation helpers, which randomises the
      ordering of candidate resources in order to explore different layouts on
      repeated runs.
    * when a flight cannot be allocated on *any* operating date the view
      delegates to ``core.services.conflict_resolution.find_alternative_slots``
      and emits a warning message containing the suggested timing shifts.  It
      also collects those suggestions and stores them in the session; a
      dedicated page is shown to the user so they can review and act on them.
    """
    flight_ids = request.POST.getlist('flight_ids')
    season_filter = request.POST.get('season', 'summer')

    if not flight_ids:
        messages.warning(request, 'No flights selected.')
        return redirect(f'/allocations/?season={season_filter}')

    from core.services.allocation import allocate_stand, allocate_gate, allocate_checkin
    from core.services.season import get_season_dates
    from core.services.conflict_resolution import find_alternative_slots

    success_count = 0
    error_count = 0
    alloc_cache = {}
    # gather flights that failed and their slot suggestions
    failure_suggestions = []

    for flight_id in flight_ids:
        try:
            flight = FlightRequest.objects.get(id=flight_id)

            # clear whatever the user had before – we are rebuilding afresh
            StandAllocation.objects.filter(flight_request=flight).delete()
            GateAllocation.objects.filter(flight_request=flight).delete()
            CheckInAllocation.objects.filter(flight_request=flight).delete()

            # compute season span
            if flight.season == 'summer':
                season_start, season_end = get_season_dates('summer', flight.year)
            else:
                season_start, season_end = get_season_dates('winter', flight.year)

            current_date = season_start
            date_success = False

            while current_date <= season_end:
                if flight.operates_on_date(current_date):
                    # randomise candidate order each run via shuffle=True
                    stand_allocated = allocate_stand(flight, current_date, alloc_cache, shuffle=True)
                    gate_allocated = allocate_gate(flight, current_date, alloc_cache, shuffle=True)
                    checkin_allocated = allocate_checkin(flight, current_date, alloc_cache, shuffle=True)

                    if stand_allocated or gate_allocated or checkin_allocated:
                        date_success = True

                current_date += timedelta(days=1)

            if date_success:
                flight.status = 'allocated'
                flight.save()
                success_count += 1
            else:
                error_count += 1
                # always warn about failure
                warn_msg = f'Flight {flight.display_flight_numbers} could not be auto-allocated.'

                # we generate recommendations only for non‑home (non-UR) flights
                suggestions = []
                if flight.airline and flight.airline.iata_code != 'UR':
                    suggestions = find_alternative_slots(flight, max_hours_search=6, interval_mins=15)
                    if suggestions:
                        txt = ', '.join(
                            f"{s['arrival_time'] or s['departure_time']}"
                            for s in suggestions
                        )
                        warn_msg += f' Try times {txt}.'
                messages.warning(request, warn_msg)

                # store for the recommendation page (serialize times to strings)
                if suggestions:
                    serialized = []
                    for s in suggestions:
                        serialized.append({
                            'arrival_time': s['arrival_time'].isoformat() if s['arrival_time'] else None,
                            'departure_time': s['departure_time'].isoformat() if s['departure_time'] else None,
                        })
                    failure_suggestions.append({
                        'flight_id': int(flight.pk),
                        'flight_display': flight.display_flight_numbers,
                        'suggestions': serialized,
                    })
        except Exception as e:
            # record the failure for debugging and skip this flight
            error_count += 1
            messages.warning(request, f'Auto-allocate error for flight_id={flight_id}: {e}')
            continue

    if success_count > 0:
        messages.success(request, f'Auto-allocated resources to {success_count} flight(s).')
    if error_count > 0 and success_count == 0:
        # if all failed we already warned per-flight, but still show aggregate
        messages.warning(request, f'Failed to allocate {error_count} flight(s).')

    # store any suggestions in the session for review
    if failure_suggestions:
        request.session['auto_recommendations'] = failure_suggestions
        messages.info(request, 'Some flights could not be placed; see recommendations.')
        return redirect('auto_recommendations')

    return redirect(f'/allocations/?season={season_filter}')


@require_POST
def season_allocations_assign(request: HttpRequest):
    """
    Assign resources to flights (bulk operation by airline or individual flights).
    """
    flight_ids = request.POST.getlist('flight_ids')
    stand_id = request.POST.get('stand_id')
    gate_id = request.POST.get('gate_id')
    counter_range = request.POST.get('counter_range')
    season_filter = request.POST.get('season', 'summer')
    
    # Convert IDs to integers (handle empty strings from "No Assignment" options)
    try:
        stand_id = int(stand_id) if stand_id and stand_id.strip() else None
    except ValueError:
        stand_id = None
    
    try:
        gate_id = int(gate_id) if gate_id and gate_id.strip() else None
    except ValueError:
        gate_id = None
    
    if not flight_ids:
        messages.warning(request, 'No flights selected.')
        return redirect(f'/allocations/?season={season_filter}')
    
    try:
        assigned_count = 0
        
        for flight_id in flight_ids:
            try:
                flight = FlightRequest.objects.get(id=flight_id)
            except:
                continue
            
            resource_assigned = False
            
            # Assign stand (apply to all dates where flight operates)
            if stand_id:
                try:
                    stand = ParkingStand.objects.get(id=stand_id)
                    from core.services.season import get_season_dates
                    
                    if flight.season == 'summer':
                        season_start, season_end = get_season_dates('summer', flight.year)
                    else:
                        season_start, season_end = get_season_dates('winter', flight.year)
                    
                    # Create allocation for each date the flight operates
                    current_date = season_start
                    while current_date <= season_end:
                        if flight.operates_on_date(current_date):
                            arrival = flight.arrival_time or flight.departure_time
                            departure = flight.departure_time or flight.arrival_time
                            if arrival and departure:
                                StandAllocation.objects.update_or_create(
                                    flight_request=flight,
                                    date=current_date,
                                    defaults={
                                        'stand': stand,
                                        'start_time': arrival,
                                        'end_time': departure,
                                    }
                                )
                                resource_assigned = True
                        current_date += timedelta(days=1)
                except Exception as e:
                    messages.warning(request, f'Error assigning stand: {str(e)}')
            
            # Assign gate (apply to all dates where flight operates)
            if gate_id and flight.operation_type != 'arrival':
                try:
                    gate = Gate.objects.get(id=gate_id)
                    from core.services.season import get_season_dates
                    
                    if flight.season == 'summer':
                        season_start, season_end = get_season_dates('summer', flight.year)
                    else:
                        season_start, season_end = get_season_dates('winter', flight.year)
                    
                    current_date = season_start
                    while current_date <= season_end:
                        if flight.operates_on_date(current_date):
                            departure_time = flight.departure_time
                            if departure_time:
                                gate_close = time_subtract_minutes(departure_time, 15)
                                gate_open = time_subtract_minutes(departure_time, 45)
                                GateAllocation.objects.update_or_create(
                                    flight_request=flight,
                                    date=current_date,
                                    defaults={
                                        'gate': gate,
                                        'start_time': gate_open,
                                        'end_time': gate_close,
                                    }
                                )
                                resource_assigned = True
                        current_date += timedelta(days=1)
                except Exception as e:
                    messages.warning(request, f'Error assigning gate: {str(e)}')
            
            # Assign check-in counters (apply to all dates where flight operates)
            if counter_range and flight.operation_type != 'arrival':
                try:
                    counter_from, counter_to = map(int, counter_range.split('-'))
                    from core.services.season import get_season_dates
                    
                    if flight.season == 'summer':
                        season_start, season_end = get_season_dates('summer', flight.year)
                    else:
                        season_start, season_end = get_season_dates('winter', flight.year)
                    
                    current_date = season_start
                    while current_date <= season_end:
                        if flight.operates_on_date(current_date):
                            departure_time = flight.departure_time
                            if departure_time:
                                checkin_close = time_subtract_minutes(departure_time, 60)
                                checkin_open = time_subtract_minutes(checkin_close, flight.checkin_duration_hours * 60)
                                CheckInAllocation.objects.update_or_create(
                                    flight_request=flight,
                                    date=current_date,
                                    defaults={
                                        'counter_from': counter_from,
                                        'counter_to': counter_to,
                                        'start_time': checkin_open,
                                        'end_time': checkin_close,
                                    }
                                )
                                resource_assigned = True
                        current_date += timedelta(days=1)
                except Exception as e:
                    messages.warning(request, f'Error assigning counters: {str(e)}')
            
            # Update flight status if any resource was assigned
            if resource_assigned:
                flight.status = 'allocated'
                flight.save(update_fields=['status', 'updated_at'])
                assigned_count += 1
        
        messages.success(request, f'Resources assigned to {assigned_count} flight(s).')
    except Exception as e:
        messages.error(request, f'Error during bulk assignment: {str(e)}')
    
    return redirect(f'/allocations/?season={season_filter}')


# ─── Resources ────────────────────────────────────────────────────────────────

def resources_view(request: HttpRequest):
    apron1_stands = ParkingStand.objects.filter(apron='apron1', parent_stand__isnull=True, is_active=True).prefetch_related('sub_stands', 'gates')
    apron1ext_stands = ParkingStand.objects.filter(apron='apron1ext', is_active=True)
    gates = Gate.objects.filter(is_active=True).select_related('connected_stand')
    counters = CheckInCounter.objects.filter(is_active=True)

    context = {
        'apron1_stands': apron1_stands,
        'apron1ext_stands': apron1ext_stands,
        'gates': gates,
        'counters': counters,
        'counter_total': counters.count(),
        'dedicated_counters': counters.filter(is_dedicated_home_airline=True),
        'common_counters': counters.filter(is_dedicated_home_airline=False),
        'active_page': 'resources',
    }
    return render(request, 'resources.html', context)


# recommendations page (see season_allocations_auto for population)
def auto_recommendations(request: HttpRequest):
    """
    Display recommendations stored in session by the bulk auto-allocation
    routine.  The page is shown immediately after that action when flights
    could not be placed; the session key is then cleared so the list is only
    visible once.
    """
    recs = request.session.pop('auto_recommendations', [])
    return render(request, 'auto_recommendations.html', {
        'recommendations': recs,
        'active_page': 'allocations',
    })


# ─── Reference Data Management ────────────────────────────────────────────────

# Airlines Management
def airlines_list(request: HttpRequest):
    airlines = Airline.objects.all().order_by('name')
    context = {
        'airlines': airlines,
        'active_page': 'airlines',
    }
    return render(request, 'admin/airlines_list.html', context)


def airline_new(request: HttpRequest):
    context = {
        'active_page': 'airlines',
        'form_action': 'new',
    }
    return render(request, 'admin/airline_form.html', context)


def airline_create(request: HttpRequest):
    if request.method != 'POST':
        return redirect('airlines_list')
    
    try:
        Airline.objects.create(
            iata_code=request.POST.get('iata_code', '').strip().upper(),
            icao_code=request.POST.get('icao_code', '').strip().upper(),
            name=request.POST.get('name', '').strip(),
            is_home_airline=request.POST.get('is_home_airline') == 'on',
        )
        messages.success(request, 'Airline added successfully.')
        return redirect('airlines_list')
    except Exception as e:
        messages.error(request, f'Error adding airline: {str(e)}')
        return redirect('airline_new')


def airline_edit(request: HttpRequest, pk: int):
    airline = get_object_or_404(Airline, pk=pk)
    context = {
        'airline': airline,
        'active_page': 'airlines',
        'form_action': 'edit',
    }
    return render(request, 'admin/airline_form.html', context)


def airline_update(request: HttpRequest, pk: int):
    if request.method != 'POST':
        return redirect('airline_edit', pk=pk)
    
    airline = get_object_or_404(Airline, pk=pk)
    try:
        airline.iata_code = request.POST.get('iata_code', '').strip().upper()
        airline.icao_code = request.POST.get('icao_code', '').strip().upper()
        airline.name = request.POST.get('name', '').strip()
        airline.is_home_airline = request.POST.get('is_home_airline') == 'on'
        airline.save()
        messages.success(request, 'Airline updated successfully.')
        return redirect('airlines_list')
    except Exception as e:
        messages.error(request, f'Error updating airline: {str(e)}')
        return redirect('airline_edit', pk=pk)


@require_POST
def airline_delete(request: HttpRequest, pk: int):
    airline = get_object_or_404(Airline, pk=pk)
    airline.delete()
    messages.success(request, 'Airline deleted.')
    return redirect('airlines_list')


# Airports Management
def airports_list(request: HttpRequest):
    airports = Airport.objects.all().order_by('iata_code')
    context = {
        'airports': airports,
        'active_page': 'airports',
    }
    return render(request, 'admin/airports_list.html', context)


def airport_new(request: HttpRequest):
    context = {
        'active_page': 'airports',
        'form_action': 'new',
    }
    return render(request, 'admin/airport_form.html', context)


def airport_create(request: HttpRequest):
    if request.method != 'POST':
        return redirect('airports_list')
    
    try:
        Airport.objects.create(
            iata_code=request.POST.get('iata_code', '').strip().upper(),
            icao_code=request.POST.get('icao_code', '').strip().upper() or None,
            city_name=request.POST.get('city_name', '').strip(),
            country=request.POST.get('country', '').strip(),
        )
        messages.success(request, 'Airport added successfully.')
        return redirect('airports_list')
    except Exception as e:
        messages.error(request, f'Error adding airport: {str(e)}')
        return redirect('airport_new')


def airport_edit(request: HttpRequest, pk: int):
    airport = get_object_or_404(Airport, pk=pk)
    context = {
        'airport': airport,
        'active_page': 'airports',
        'form_action': 'edit',
    }
    return render(request, 'admin/airport_form.html', context)


def airport_update(request: HttpRequest, pk: int):
    if request.method != 'POST':
        return redirect('airport_edit', pk=pk)
    
    airport = get_object_or_404(Airport, pk=pk)
    try:
        airport.iata_code = request.POST.get('iata_code', '').strip().upper()
        airport.icao_code = request.POST.get('icao_code', '').strip().upper() or None
        airport.city_name = request.POST.get('city_name', '').strip()
        airport.country = request.POST.get('country', '').strip()
        airport.save()
        messages.success(request, 'Airport updated successfully.')
        return redirect('airports_list')
    except Exception as e:
        messages.error(request, f'Error updating airport: {str(e)}')
        return redirect('airport_edit', pk=pk)


@require_POST
def airport_delete(request: HttpRequest, pk: int):
    airport = get_object_or_404(Airport, pk=pk)
    airport.delete()
    messages.success(request, 'Airport deleted.')
    return redirect('airports_list')


# Aircraft Types Management
def aircraft_types_list(request: HttpRequest):
    aircraft_types = AircraftType.objects.all().order_by('code')
    context = {
        'aircraft_types': aircraft_types,
        'active_page': 'aircraft_types',
    }
    return render(request, 'admin/aircraft_types_list.html', context)


def aircraft_type_new(request: HttpRequest):
    context = {
        'size_codes': ['A', 'B', 'C', 'D', 'E', 'F'],
        'categories': ['wide_body', 'narrow_body', 'regional', 'turboprop'],
        'active_page': 'aircraft_types',
        'form_action': 'new',
    }
    return render(request, 'admin/aircraft_type_form.html', context)


def aircraft_type_create(request: HttpRequest):
    if request.method != 'POST':
        return redirect('aircraft_types_list')
    
    try:
        AircraftType.objects.create(
            code=request.POST.get('code', '').strip().upper(),
            name=request.POST.get('name', '').strip(),
            manufacturer=request.POST.get('manufacturer', '').strip(),
            category=request.POST.get('category', ''),
            size_code=request.POST.get('size_code', ''),
            is_wide_body=request.POST.get('is_wide_body') == 'on',
            pax_capacity=int(request.POST.get('pax_capacity', 0)) if request.POST.get('pax_capacity') else 0,
        )
        messages.success(request, 'Aircraft type added successfully.')
        return redirect('aircraft_types_list')
    except Exception as e:
        messages.error(request, f'Error adding aircraft type: {str(e)}')
        return redirect('aircraft_type_new')


def aircraft_type_edit(request: HttpRequest, pk: int):
    aircraft_type = get_object_or_404(AircraftType, pk=pk)
    context = {
        'aircraft_type': aircraft_type,
        'size_codes': ['A', 'B', 'C', 'D', 'E', 'F'],
        'categories': ['wide_body', 'narrow_body', 'regional', 'turboprop'],
        'active_page': 'aircraft_types',
        'form_action': 'edit',
    }
    return render(request, 'admin/aircraft_type_form.html', context)


def aircraft_type_update(request: HttpRequest, pk: int):
    if request.method != 'POST':
        return redirect('aircraft_type_edit', pk=pk)
    
    aircraft_type = get_object_or_404(AircraftType, pk=pk)
    try:
        aircraft_type.code = request.POST.get('code', '').strip().upper()
        aircraft_type.name = request.POST.get('name', '').strip()
        aircraft_type.manufacturer = request.POST.get('manufacturer', '').strip()
        aircraft_type.category = request.POST.get('category', '')
        aircraft_type.size_code = request.POST.get('size_code', '')
        aircraft_type.is_wide_body = request.POST.get('is_wide_body') == 'on'
        aircraft_type.pax_capacity = int(request.POST.get('pax_capacity', 0)) if request.POST.get('pax_capacity') else 0
        aircraft_type.save()
        messages.success(request, 'Aircraft type updated successfully.')
        return redirect('aircraft_types_list')
    except Exception as e:
        messages.error(request, f'Error updating aircraft type: {str(e)}')
        return redirect('aircraft_type_edit', pk=pk)


@require_POST
def aircraft_type_delete(request: HttpRequest, pk: int):
    aircraft_type = get_object_or_404(AircraftType, pk=pk)
    aircraft_type.delete()
    messages.success(request, 'Aircraft type deleted.')
    return redirect('aircraft_types_list')


# ─── Reports ──────────────────────────────────────────────────────────────────

from django.http import HttpResponse
from core.services.reports import generate_report_data, generate_daily_analysis
from core.services.season import get_current_season

def reports_dashboard(request: HttpRequest):
    season, year, _, _ = get_current_season()
    season = request.GET.get('season', season)
    year = int(request.GET.get('year', year))
    return render(request, 'reports/index.html', {
        'season': season,
        'year': year,
        'active_page': 'reports',
    })


def _build_season_schedule_rows(season: str, year: int) -> list[dict[str, Any]]:
    """Build the list of row dicts for the season schedule table / Excel export."""
    from core.services.season import get_season_dates
    season_start, season_end = get_season_dates(season, year)

    flights = (
        FlightRequest.objects
        .filter(season=season, year=year)
        .select_related('airline', 'aircraft_type', 'origin', 'destination')
        .order_by('airline__iata_code', 'arrival_flight_number', 'departure_flight_number')
    )

    day_order = ['sunday', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday']
    rows = []

    for flight in flights:
        start_date = flight.valid_from or season_start
        end_date   = flight.valid_to   or season_end

        days = {
            day: ('Y' if (flight.days_of_operation & DAY_MASK[day]) else '-')
            for day in day_order
        }

        is_home = flight.airline.is_home_airline

        # ── Arrival row ──────────────────────────────────────────────────────
        if flight.operation_type in ('arrival', 'turnaround'):
            rows.append({
                'arr_dep':       'A',
                'start_date':    start_date.strftime('%Y.%m.%d'),
                'end_date':      end_date.strftime('%Y.%m.%d'),
                'operator':      flight.airline.iata_code,
                'flight_no':     flight.arrival_flight_number,
                'acft_type':     flight.aircraft_type.code,
                'time':          flight.arrival_time.strftime('%H:%M') if flight.arrival_time else '',
                'from_iata':     flight.origin.iata_code if flight.origin else '',
                'from_icao':     (flight.origin.icao_code or '') if flight.origin else '',
                'days':          days,
                'irregular':     '',
                'dom_int':       'I',
                'link_flight':   flight.departure_flight_number if not is_home else '',
                'link_day':      flight.ground_days,
                'codeshare':     '',
                'master_flight': '',
                'service_type':  'J',
                'remark':        flight.airline.name,
            })

        # ── Departure row ────────────────────────────────────────────────────
        if flight.operation_type in ('departure', 'turnaround'):
            rows.append({
                'arr_dep':       'D',
                'start_date':    start_date.strftime('%Y.%m.%d'),
                'end_date':      end_date.strftime('%Y.%m.%d'),
                'operator':      flight.airline.iata_code,
                'flight_no':     flight.departure_flight_number,
                'acft_type':     flight.aircraft_type.code,
                'time':          flight.departure_time.strftime('%H:%M') if flight.departure_time else '',
                'from_iata':     flight.destination.iata_code if flight.destination else '',
                'from_icao':     (flight.destination.icao_code or '') if flight.destination else '',
                'days':          days,
                'irregular':     '',
                'dom_int':       'I',
                'link_flight':   flight.arrival_flight_number if not is_home else '',
                'link_day':      flight.ground_days,
                'codeshare':     '',
                'master_flight': '',
                'service_type':  'J',
                'remark':        flight.airline.name,
            })

    return rows


def season_schedule_view(request: HttpRequest):
    season, year, _, _ = get_current_season()
    season = request.GET.get('season', season)
    year   = int(request.GET.get('year', year))

    available_seasons = list(
        FlightRequest.objects
        .order_by('season', 'year')
        .values('season', 'year')
        .distinct()
    )

    rows = _build_season_schedule_rows(season, year)

    return render(request, 'reports/season_schedule.html', {
        'rows':              rows,
        'season':            season,
        'year':              year,
        'available_seasons': available_seasons,
        'active_page':       'reports',
    })


def season_schedule_export_excel(request: HttpRequest):
    """Export in SEASON-IMPORT column format while keeping A/D rows separate."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.services.season import get_season_dates

    season = request.GET.get('season', 'summer')
    year   = int(request.GET.get('year', 2026))
    rows = _build_season_schedule_rows(season, year)

    season_start, season_end = get_season_dates(season, year)
    season_code = ('S' if season == 'summer' else 'W') + str(year)

    airlines_by_iata = {
        a.iata_code: a for a in Airline.objects.all().only('iata_code', 'icao_code')
    }

    def to_yyyymmdd(v: str | None) -> str:
        # rows currently carry dates as YYYY.MM.DD
        if not v:
            return ''
        return str(v).replace('.', '')

    def to_hhmm_with_space(v: str | None) -> str:
        # rows currently carry times as HH:MM
        if not v:
            return ''
        return str(v).replace(':', ' ')

    def flight_with_space(operator_iata: str | None, flight_no: str | None) -> str:
        no = (flight_no or '').strip()
        if not no:
            return ''
        op = (operator_iata or '').strip().upper()
        if op and no.upper().startswith(op):
            return f"{op} {no[len(op):]}"
        return no

    def acft_iata_code(icao_code: str | None) -> str:
        code = (icao_code or '').strip()
        if not code:
            return ''
        # Example: B788 -> 788, A332 -> 332, DH8C -> H8C
        return code[1:] if len(code) > 1 else code

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = season_code

    # Row 1: short template column codes (exact order from SEASON-IMPORT)
    ws.append([
        'ARP', 'TER', 'SSC', 'FSD', 'FED', 'FLT', 'FLO', 'ORG', 'STA',
        'TYA', 'TYS', 'TYC', 'DES', 'STD', 'LKT', 'LKD',
        'SUN', 'MON', 'TUE', 'WED', 'THU', 'FRI', 'SAT',
        'IRR', 'FST', 'TRA', 'RMK',
    ])

    # Row 2: long template labels (exact wording from sample file)
    ws.append([
        '*Service Airport \nIATA Code',
        '*Terminal No',
        '*Season Code',
        '*Season \nStart Date\n(YYYYMMDD)',
        '*Season \nEnd Date\n(YYYYMMDD)',
        '*Flight IATA Code\n(AA 123)\n',
        'Flight Airline ICAO Code',
        '*Origin\nAirport \nIATA code',
        '*Scheduled \nArrival Time\n(hhmm)',
        '*Aircraft IATA Type code\n',
        '*Aircraft IATA Subtype code\n',
        '*Aircraft ICAO Type code\n',
        'Destination\nAirport \nIATA code\n(Required if there is a link flight)',
        'Scheduled\nDeparture Time\n(hhmm)\n(Required if there is a link flight)',
        'Link Flight IATA Code\n(AA 123)\n(Required if there is a link flight)',
        'Link Days',
        '*Y/N', '*Y/N', '*Y/N', '*Y/N', '*Y/N', '*Y/N', '*Y/N',
        '*Y/N',
        '*Service Type\n(A~Z)',
        'Transfer \nY/N',
        'Remark',
    ])

    thin = Border(
        left=Side(style='thin', color='B0C8D4'),
        right=Side(style='thin', color='B0C8D4'),
        top=Side(style='thin', color='B0C8D4'),
        bottom=Side(style='thin', color='B0C8D4'),
    )
    for cell in ws[1]:
        cell.fill = PatternFill('solid', fgColor='1A6E8E')
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin
    ws.row_dimensions[1].height = 20

    for cell in ws[2]:
        cell.fill = PatternFill('solid', fgColor='2C8FAA')
        cell.font = Font(bold=False, color='FFFFFF', size=7)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[2].height = 52

    fill_odd  = PatternFill('solid', fgColor='E8F4F8')
    fill_even = PatternFill('solid', fgColor='FFFFFF')
    for idx, row in enumerate(rows, start=3):
        operator = row['operator']
        airline = airlines_by_iata.get(operator)
        flo = airline.icao_code if airline else ''

        start_val = to_yyyymmdd(row['start_date']) or season_start.strftime('%Y%m%d')
        end_val = to_yyyymmdd(row['end_date']) or season_end.strftime('%Y%m%d')

        acft_icao = row['acft_type']
        acft_iata = acft_iata_code(acft_icao)

        ws.append([
            'EBB',
            'T01',
            season_code,
            start_val,
            end_val,
            flight_with_space(operator, row['flight_no']),
            flo,
            row['from_iata'],
            to_hhmm_with_space(row['time']),
            acft_iata,
            acft_iata,
            acft_icao,
            row['from_iata'] if row['arr_dep'] == 'D' else row['from_iata'],
            to_hhmm_with_space(row['time']) if row['arr_dep'] == 'D' else '',
            flight_with_space(operator, row['link_flight']),
            row['link_day'],
            row['days']['sunday'],
            row['days']['monday'],
            row['days']['tuesday'],
            row['days']['wednesday'],
            row['days']['thursday'],
            row['days']['friday'],
            row['days']['saturday'],
            row['irregular'],
            row['service_type'],
            row['codeshare'],
            row['remark'],
        ])

        fill = fill_odd if idx % 2 == 0 else fill_even
        for col_i, cell in enumerate(ws[idx], start=1):
            cell.fill  = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9, bold=True) if col_i in (6, 7) else Font(size=9)

        ws.row_dimensions[idx].height = 15

    for i, w in enumerate(
        [7, 6, 8, 11, 11, 11, 7, 7, 9, 7, 7, 9, 9, 9, 12, 8, 5, 5, 5, 5, 5, 5, 5, 6, 6, 6, 22],
        start=1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A3'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="SEASON-IMPORT_{season_code}.xlsx"'
    wb.save(response)
    return response

def daily_analysis(request: HttpRequest):
    season, year, _, _ = get_current_season()
    season = request.GET.get('season', season)
    year = int(request.GET.get('year', year))
    day_of_week = request.GET.get('day', 'monday').lower()

    # delegate heavy lifting to reports service
    data = generate_daily_analysis(season, year, day_of_week)
    return render(request, 'reports/daily_analysis.html', data)

def reports_export_excel(request: HttpRequest):
    import openpyxl
    season = request.GET.get('season', 'summer')
    year = int(request.GET.get('year', date.today().year))
    
    data = generate_report_data(season, year)
    
    wb = openpyxl.Workbook()
    # Sheet 1: Stats
    ws = wb.active
    assert ws is not None
    ws.title = "Flight Statistics"
    ws.append(["Season", season.capitalize(), "Year", year])
    ws.append(["Total Flights", data.get('total_flights')])
    ws.append([])
    ws.append(["Operation Type", "Count"])
    for op, c in data.get('ops_dict', {}).items():
        ws.append([op, c])
        
    ws.append([])
    ws.append(["Aircraft Size", "Count"])
    for size, c in data.get('size_dict', {}).items():
        ws.append([f"Code {size}", c])
        
    vs = wb.create_sheet("Airline Performance")
    vs.append(["Airline", "Code", "Flights", "Estimated Weekly Pax", "Top Route"])
    for p in data.get('airline_perf', []):
        vs.append([p['airline'], p['code'], p['flights'], p['weekly_pax_est'], p['top_route']])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=reports_{season}_{year}.xlsx'
    wb.save(response)
    return response

def reports_export_pdf(request: HttpRequest):
    season = request.GET.get('season', 'summer')
    year = int(request.GET.get('year', date.today().year))
    
    data = generate_report_data(season, year)
    data['is_print'] = True
    return render(request, 'reports/pdf.html', data)


